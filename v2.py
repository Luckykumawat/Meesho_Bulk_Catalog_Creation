import os
import shutil
from openpyxl import load_workbook

# ---------- CONFIG ----------
MASTER_FILE = "master.xlsx"    # master list with Model Name in column A
TEMPLATE_FILE = "template.xlsx"     # template workbook to copy
OUTPUT_DIR = "output"               # folder where generated files will be saved
PROCESSED_LOG = "processed_models.txt"  # file that stores names of already processed models

TARGET_SHEET_NAME = "Mobile-Cases---Covers-Fill this"  # target sheet name in template

os.makedirs(OUTPUT_DIR, exist_ok=True)

# ---------------- HELPER FUNCTIONS ---------------- #

def get_top_left_of_merge(ws):
    merged_map = {}
    for merged_range in ws.merged_cells.ranges:
        min_row = merged_range.min_row
        min_col = merged_range.min_col
        for r in range(merged_range.min_row, merged_range.max_row + 1):
            for c in range(merged_range.min_col, merged_range.max_col + 1):
                merged_map[(r, c)] = (min_row, min_col)
    return merged_map


def safe_set(ws, row, col, value, merged_map):
    key = (row, col)
    if key in merged_map:
        row, col = merged_map[key]
    ws.cell(row=row, column=col, value=value)


def fill_template_for_model(model_name, output_file):
    wb = load_workbook(output_file)
    if TARGET_SHEET_NAME in wb.sheetnames:
        ws = wb[TARGET_SHEET_NAME]
    elif len(wb.worksheets) > 1:
        ws = wb.worksheets[1]
    else:
        ws = wb.active

    merged_map = get_top_left_of_merge(ws)

    # D=4, X=24, AJ=36, AK=37
    for row in range(5, 105):
        safe_set(ws, row, 4, f"{model_name} / Sticker Printed Back Cover", merged_map)
        safe_set(ws, row, 24, model_name, merged_map)
        safe_set(ws, row, 36, f"Sticker {model_name} EG {row - 4}", merged_map)
        safe_set(ws, row, 37, f"Sticker {model_name} EG {row - 4}", merged_map)

    wb.save(output_file)
    wb.close()


def read_processed_models():
    if not os.path.exists(PROCESSED_LOG):
        return set()
    with open(PROCESSED_LOG, "r", encoding="utf-8") as f:
        return set(line.strip() for line in f if line.strip())


def save_processed_model(model_name):
    with open(PROCESSED_LOG, "a", encoding="utf-8") as f:
        f.write(model_name + "\n")


# ---------------- MAIN LOGIC ---------------- #

def main():
    if not os.path.exists(MASTER_FILE):
        print(f"❌ Master file '{MASTER_FILE}' not found.")
        return
    if not os.path.exists(TEMPLATE_FILE):
        print(f"❌ Template file '{TEMPLATE_FILE}' not found.")
        return

    processed = read_processed_models()

    master_wb = load_workbook(MASTER_FILE, read_only=True)
    master_ws = master_wb.active

    models = []
    for row in master_ws.iter_rows(min_col=1, max_col=1, min_row=2, values_only=True):
        value = row[0]
        if value and str(value).strip():
            models.append(str(value).strip())

    master_wb.close()

    print(f"🟢 Found {len(models)} models in master file.")
    new_models = [m for m in models if m not in processed]

    if not new_models:
        print("✅ No new models to process. Everything is up-to-date.")
        return

    print(f"✨ Generating {len(new_models)} new files...")

    for model in new_models:
        safe_name = "".join(c for c in model if c.isalnum() or c in (" ", "_", "-")).strip()
        output_path = os.path.join(OUTPUT_DIR, f"{safe_name}.xlsx")

        shutil.copy(TEMPLATE_FILE, output_path)

        try:
            fill_template_for_model(safe_name, output_path)
            save_processed_model(model)
            print(f"✅ Created & updated: {safe_name}.xlsx")
        except Exception as e:
            print(f"❌ Failed for '{safe_name}': {e}")

    print("\n🎉 All new models processed successfully!")


if __name__ == "__main__":
    main()
